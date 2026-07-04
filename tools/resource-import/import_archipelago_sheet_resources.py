#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import requests


ADMISSIBLE_STABILITY = {"stable", "unstable"}
GITHUB_RE = re.compile(r"^https?://github\.com/([^/]+)/([^/#?]+)(?:/([^#?]*))?", re.I)
URL_RE = re.compile(r"https?://[^\s<>\")]+", re.I)


@dataclass
class SheetResource:
    game: str
    sheet: str
    row: int
    status: str
    pr_status: str = ""
    link_label: str = ""
    url: str = ""
    notes: str = ""
    adult: bool = False
    source_kind: str = "cell"


@dataclass
class DownloadAsset:
    game: str
    status: str
    kind: str
    url: str
    source_url: str
    source_label: str
    file_name: str
    path: str = ""
    size: int = 0
    skipped: str = ""
    error: str = ""


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "unknown"


def clean_url(url: str) -> str:
    url = html.unescape(str(url or "").strip())
    return url.rstrip(".,);")


def cell_urls(cell: Any) -> list[tuple[str, str]]:
    found: list[tuple[str, str]] = []
    if cell.hyperlink and cell.hyperlink.target:
        found.append((str(cell.value or ""), clean_url(cell.hyperlink.target)))
    if isinstance(cell.value, str):
        for url in URL_RE.findall(cell.value):
            found.append((str(cell.value or ""), clean_url(url)))
    if cell.comment and cell.comment.text:
        text = str(cell.comment.text)
        for url in URL_RE.findall(text):
            found.append(("comment", clean_url(url)))
    dedup: list[tuple[str, str]] = []
    seen: set[str] = set()
    for label, url in found:
        if url and url not in seen:
            seen.add(url)
            dedup.append((label, url))
    return dedup


def load_inventory(xlsx_path: Path) -> list[SheetResource]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    resources: list[SheetResource] = []

    if "Playable Worlds" in wb.sheetnames:
        ws = wb["Playable Worlds"]
        for row_idx in range(7, ws.max_row + 1):
            game = str(ws.cell(row_idx, 1).value or "").strip()
            stability = str(ws.cell(row_idx, 2).value or "").strip()
            pr_status = str(ws.cell(row_idx, 3).value or "").strip()
            link_label = str(ws.cell(row_idx, 4).value or "").strip()
            adult = bool(ws.cell(row_idx, 5).value)
            notes = str(ws.cell(row_idx, 6).value or "").strip()
            if not game or stability.lower() not in ADMISSIBLE_STABILITY or adult:
                continue
            for label, url in cell_urls(ws.cell(row_idx, 4)):
                resources.append(SheetResource(
                    game=game,
                    sheet=ws.title,
                    row=row_idx,
                    status=stability,
                    pr_status=pr_status,
                    link_label=link_label or label,
                    url=url,
                    notes=notes,
                    adult=adult,
                    source_kind="cell-or-comment",
                ))

    if "Core-Verified Worlds" in wb.sheetnames:
        ws = wb["Core-Verified Worlds"]
        for row_idx in range(3, ws.max_row + 1):
            game = str(ws.cell(row_idx, 1).value or "").strip()
            if not game:
                continue
            for col in range(2, 5):
                label, url = "", ""
                urls = cell_urls(ws.cell(row_idx, col))
                for label, url in urls:
                    resources.append(SheetResource(
                        game=game,
                        sheet=ws.title,
                        row=row_idx,
                        status="Core-Verified",
                        link_label=str(ws.cell(2, col).value or label),
                        url=url,
                        source_kind="core-link",
                    ))

    return resources


class GitHubClient:
    def __init__(self, timeout: int = 30) -> None:
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/vnd.github+json",
            "User-Agent": "SekaiLink-resource-importer",
        })
        token = os.environ.get("GITHUB_TOKEN") or os.environ.get("GH_TOKEN")
        if not token:
            try:
                token = subprocess.check_output(["gh", "auth", "token"], text=True, stderr=subprocess.DEVNULL).strip()
            except Exception:
                token = ""
        if token:
            self.session.headers["Authorization"] = f"Bearer {token}"
        self.timeout = timeout

    def get_json(self, url: str) -> Any:
        res = self.session.get(url, timeout=self.timeout, headers={"Accept": "application/vnd.github+json"})
        res.raise_for_status()
        return res.json()

    def get(self, url: str) -> requests.Response:
        res = self.session.get(url, timeout=self.timeout, allow_redirects=True, headers={"Accept": "text/html,application/xhtml+xml"})
        res.raise_for_status()
        return res

    def download(self, url: str, out: Path) -> int:
        with self.session.get(url, stream=True, timeout=self.timeout) as res:
            res.raise_for_status()
            out.parent.mkdir(parents=True, exist_ok=True)
            tmp = out.with_suffix(out.suffix + ".part")
            total = 0
            with tmp.open("wb") as fh:
                for chunk in res.iter_content(chunk_size=1024 * 256):
                    if not chunk:
                        continue
                    total += len(chunk)
                    fh.write(chunk)
            tmp.replace(out)
            return total


def parse_github(url: str) -> tuple[str, str, str]:
    match = GITHUB_RE.match(url)
    if not match:
        return "", "", ""
    owner, repo, path = match.group(1), match.group(2), match.group(3) or ""
    return owner, repo.removesuffix(".git"), path


def release_page_url(resource_url: str) -> str:
    owner, repo, path = parse_github(resource_url)
    if not owner:
        return ""
    if "/releases" in resource_url:
        return resource_url
    return ""


def release_api_urls(resource_url: str) -> list[str]:
    owner, repo, path = parse_github(resource_url)
    if not owner:
        return []
    path = urllib.parse.unquote(path)
    if path.startswith("releases/tag/"):
        tag = path.removeprefix("releases/tag/").split("/")[0]
        return [f"https://api.github.com/repos/{owner}/{repo}/releases/tags/{urllib.parse.quote(tag, safe='')}"]
    if path.startswith("releases/latest"):
        return [f"https://api.github.com/repos/{owner}/{repo}/releases/latest"]
    if path.startswith("releases"):
        return [f"https://api.github.com/repos/{owner}/{repo}/releases?per_page=30"]
    return []


def direct_download_url(resource_url: str) -> str:
    owner, repo, path = parse_github(resource_url)
    if not owner:
        return ""
    if "/raw/" in resource_url or "/releases/download/" in resource_url:
        return resource_url
    if path.startswith("raw/"):
        return resource_url
    return ""


def normalized_query_terms(resource_url: str) -> list[str]:
    query = urllib.parse.parse_qs(urllib.parse.urlparse(resource_url).query)
    raw = " ".join(query.get("q", []))
    raw = urllib.parse.unquote_plus(raw)
    terms = [t.lower() for t in re.findall(r"[a-z0-9]{3,}", raw, re.I)]
    ignored = {"the", "and", "ap", "expanded", "true", "world", "archipelago"}
    return [t for t in terms if t not in ignored]


def release_matches_query(release: dict[str, Any], asset_name: str, terms: list[str]) -> bool:
    if not terms:
        return True
    hay = " ".join(str(release.get(k) or "") for k in ("tag_name", "name", "body"))
    hay = f"{hay} {asset_name}".lower()
    hay_norm = re.sub(r"[^a-z0-9]+", " ", hay)
    compact = hay_norm.replace(" ", "")
    return any(term in hay_norm or term in compact for term in terms)


def classify_asset(name: str, url: str) -> str:
    lower = f"{name} {url}".lower()
    if lower.endswith(".apworld") or ".apworld" in lower:
        return "apworld"
    trackerish = "poptracker" in lower or "pop-tracker" in lower or "map-tracker" in lower
    if trackerish and re.search(r"\.(zip|yaml|json|ptpack|tar\.gz|7z)$", lower):
        return "poptracker"
    if re.search(r"(^|[-_. ])tracker([-_. ]|$)", lower) and re.search(r"\.(zip|yaml|json|ptpack|tar\.gz|7z)$", lower):
        return "tracker"
    return ""


def iter_candidate_assets(client: GitHubClient, resource: SheetResource) -> list[dict[str, Any]]:
    direct = direct_download_url(resource.url)
    if direct:
        name = Path(urllib.parse.urlparse(direct).path).name
        kind = classify_asset(name, direct)
        return [{"name": name, "download_url": direct, "kind": kind, "release": "direct"}] if kind else []
    api_urls = release_api_urls(resource.url)
    if api_urls:
        assets: list[dict[str, Any]] = []
        terms = normalized_query_terms(resource.url)
        seen_names: set[tuple[str, str]] = set()
        for api_url in api_urls:
            data = client.get_json(api_url)
            releases = data if isinstance(data, list) else [data]
            for rel in releases:
                if not isinstance(rel, dict) or rel.get("draft"):
                    continue
                for asset in rel.get("assets") or []:
                    if not isinstance(asset, dict):
                        continue
                    name = str(asset.get("name") or "")
                    dl = str(asset.get("browser_download_url") or "")
                    kind = classify_asset(name, dl)
                    if not kind or not release_matches_query(rel, name, terms):
                        continue
                    key = (kind, name.lower())
                    if key in seen_names:
                        continue
                    seen_names.add(key)
                    assets.append({
                        "name": name,
                        "download_url": dl,
                        "kind": kind,
                        "release": rel.get("tag_name") or rel.get("name") or "",
                    })
        return assets

    page_url = release_page_url(resource.url)
    if not page_url:
        return []
    assets: list[dict[str, Any]] = []
    response = client.get(page_url)
    owner, repo, _path = parse_github(response.url)
    if not owner:
        owner, repo, _path = parse_github(page_url)
    href_re = re.compile(r'href="([^"]*/releases/download/[^"]+)"')
    seen: set[str] = set()
    for href in href_re.findall(response.text):
        href = html.unescape(href)
        if href.startswith("/"):
            dl = f"https://github.com{href}"
        elif href.startswith("http"):
            dl = href
        else:
            continue
        if dl in seen:
            continue
        seen.add(dl)
        name = Path(urllib.parse.urlparse(dl).path).name
        kind = classify_asset(name, dl)
        if kind:
            assets.append({
                "name": name,
                "download_url": dl,
                "kind": kind,
                "release": "github-html",
            })
    # Source archives occasionally are the only pack distribution, but they are
    # ambiguous and can be huge, so we intentionally do not download them
    # automatically.
    return assets


def download_resources(inventory: list[SheetResource], out_dir: Path, dry_run: bool = False) -> list[DownloadAsset]:
    client = GitHubClient()
    results: list[DownloadAsset] = []
    seen_assets: set[str] = set()
    github_resources = [r for r in inventory if parse_github(r.url)[0]]
    for idx, resource in enumerate(github_resources, 1):
        print(f"[{idx}/{len(github_resources)}] {resource.game}: {resource.url}", flush=True)
        try:
            candidates = iter_candidate_assets(client, resource)
        except Exception as err:
            results.append(DownloadAsset(
                game=resource.game,
                status=resource.status,
                kind="unknown",
                url="",
                source_url=resource.url,
                source_label=resource.link_label,
                file_name="",
                error=f"{type(err).__name__}: {err}",
            ))
            continue
        if not candidates:
            results.append(DownloadAsset(
                game=resource.game,
                status=resource.status,
                kind="none",
                url="",
                source_url=resource.url,
                source_label=resource.link_label,
                file_name="",
                skipped="no_matching_release_assets",
            ))
            continue
        for candidate in candidates:
            url = candidate["download_url"]
            if not url or url in seen_assets:
                continue
            seen_assets.add(url)
            name = candidate["name"]
            kind = candidate["kind"]
            target = out_dir / kind / slugify(resource.game) / name
            result = DownloadAsset(
                game=resource.game,
                status=resource.status,
                kind=kind,
                url=url,
                source_url=resource.url,
                source_label=resource.link_label,
                file_name=name,
                path=str(target),
            )
            try:
                if dry_run:
                    result.skipped = "dry_run"
                elif target.exists() and target.stat().st_size > 0:
                    result.size = target.stat().st_size
                    result.skipped = "already_exists"
                else:
                    result.size = client.download(url, target)
                    time.sleep(0.2)
            except Exception as err:
                result.error = f"{type(err).__name__}: {err}"
            results.append(result)
    return results


def write_outputs(inventory: list[SheetResource], results: list[DownloadAsset], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "inventory.json").write_text(json.dumps([asdict(r) for r in inventory], indent=2, ensure_ascii=False), encoding="utf-8")
    (out_dir / "download-report.json").write_text(json.dumps([asdict(r) for r in results], indent=2, ensure_ascii=False), encoding="utf-8")
    downloaded = [r for r in results if r.path and not r.error and r.skipped != "dry_run"]
    errors = [r for r in results if r.error]
    skipped = [r for r in results if r.skipped]
    lines = [
        "# Archipelago Sheet Resource Import",
        "",
        "Generated from the configured Archipelago games sheet.",
        "",
        f"- Inventory rows with URLs: {len(inventory)}",
        f"- Downloaded/existing matching assets: {len(downloaded)}",
        f"- Skipped records: {len(skipped)}",
        f"- Errors: {len(errors)}",
        "",
        "Downloaded assets are quarantined by kind under this folder and are not automatically enabled in the runtime.",
        "",
        "## Downloaded Assets",
        "",
    ]
    for r in downloaded:
        lines.append(f"- `{r.kind}` `{r.game}`: `{r.file_name}`")
    if errors:
        lines.extend(["", "## Errors", ""])
        for r in errors[:200]:
            lines.append(f"- `{r.game}` from {r.source_url}: {r.error}")
    (out_dir / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    repo_root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=os.environ.get("SEKAILINK_ARCHIPELAGO_SHEET", "Archipelago Games Sheet.xlsx"))
    parser.add_argument("--out", default=str(repo_root / "runtime/downloaded-resources/archipelago-sheet"))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    out = Path(args.out)
    if not xlsx.exists():
        print(f"missing xlsx: {xlsx}", file=sys.stderr)
        return 2
    inventory = load_inventory(xlsx)
    results = download_resources(inventory, out, dry_run=args.dry_run)
    write_outputs(inventory, results, out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
